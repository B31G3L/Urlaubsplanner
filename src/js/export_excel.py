#!/usr/bin/env python3
"""
Excel Export für Teamplanner
Erstellt professionelle Excel-Datei mit Mitarbeiter-Statistiken
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def format_number(value):
    """Formatiert Zahlen: Ganzzahlen ohne Nachkommastellen, Dezimalzahlen mit 2 Stellen"""
    if value is None or value == '':
        return 0
    
    try:
        num = float(value)
        if num == int(num):
            return int(num)
        return round(num, 2)
    except (ValueError, TypeError):
        return 0

def create_excel(data, jahr, output_path):
    """Erstellt Excel-Datei mit Mitarbeiter-Daten"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Urlaubsplaner {jahr}"
    
    # Header-Style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    headers = [
        "Vorname", "Nachname", "Abteilung", "Anspruch", "Übertrag", 
        "Verfügbar", "Genommen", "Rest", "Krank", "Schulung", "Überstunden"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Daten
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    for row_idx, mitarbeiter in enumerate(data, start=2):
        # Vorname
        cell = ws.cell(row=row_idx, column=1, value=mitarbeiter.get('vorname', ''))
        cell.alignment = left_alignment
        cell.border = border
        
        # Nachname
        cell = ws.cell(row=row_idx, column=2, value=mitarbeiter.get('nachname', ''))
        cell.alignment = left_alignment
        cell.border = border
        
        # Abteilung
        cell = ws.cell(row=row_idx, column=3, value=mitarbeiter.get('abteilung', ''))
        cell.alignment = center_alignment
        cell.border = border
        
        # Anspruch
        cell = ws.cell(row=row_idx, column=4, value=format_number(mitarbeiter.get('urlaub_anspruch', 0)))
        cell.alignment = center_alignment
        cell.border = border
        
        # Übertrag
        cell = ws.cell(row=row_idx, column=5, value=format_number(mitarbeiter.get('urlaub_uebertrag', 0)))
        cell.alignment = center_alignment
        cell.border = border
        
        # Verfügbar (Formel)
        cell = ws.cell(row=row_idx, column=6, value=f"=D{row_idx}+E{row_idx}")
        cell.alignment = center_alignment
        cell.border = border
        
        # Genommen
        cell = ws.cell(row=row_idx, column=7, value=format_number(mitarbeiter.get('urlaub_genommen', 0)))
        cell.alignment = center_alignment
        cell.border = border
        
        # Rest (Formel)
        cell = ws.cell(row=row_idx, column=8, value=f"=F{row_idx}-G{row_idx}")
        cell.alignment = center_alignment
        cell.border = border
        
        # Krank
        cell = ws.cell(row=row_idx, column=9, value=format_number(mitarbeiter.get('krankheit', 0)))
        cell.alignment = center_alignment
        cell.border = border
        
        # Schulung
        cell = ws.cell(row=row_idx, column=10, value=format_number(mitarbeiter.get('schulung', 0)))
        cell.alignment = center_alignment
        cell.border = border
        
        # Überstunden
        cell = ws.cell(row=row_idx, column=11, value=format_number(mitarbeiter.get('ueberstunden', 0)))
        cell.alignment = center_alignment
        cell.border = border
    
    # Spaltenbreiten
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 14
    
    # Freeze Panes (Header fixieren)
    ws.freeze_panes = 'A2'
    
    # Speichern
    wb.save(output_path)

if __name__ == "__main__":
    try:
        if len(sys.argv) != 4:
            print(json.dumps({
                "success": False,
                "error": "Usage: export_excel.py <json_file_path> <jahr> <output_path>"
            }))
            sys.exit(1)
        
        json_file_path = sys.argv[1]
        jahr = sys.argv[2]
        output_path = sys.argv[3]
        
        # Lese JSON aus Datei statt aus Command-Line
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        create_excel(data, jahr, output_path)
        
        print(json.dumps({
            "success": True,
            "path": output_path
        }))
        
    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        sys.exit(1)