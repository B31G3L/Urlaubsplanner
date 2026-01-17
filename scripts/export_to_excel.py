#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel-Export für Teamplanner
Erstellt eine formatierte Excel-Datei aus Urlaubsdaten
"""

import sys
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("FEHLER: openpyxl nicht installiert!", file=sys.stderr)
    print("Installiere mit: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def create_excel(data, output_path):
    """Erstellt Excel-Datei mit formatierten Urlaubsdaten"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Urlaubsübersicht"
    
    # Header-Style
    header_fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header schreiben
    headers = ["Mitarbeiter", "Abteilung", "Anspruch", "Übertrag", "Verfügbar", "Genommen", "Rest", "Krank", "Schulung", "Überstunden"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Daten schreiben
    for row_idx, entry in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=entry.get('mitarbeiter', ''))
        ws.cell(row=row_idx, column=2, value=entry.get('abteilung', ''))
        ws.cell(row=row_idx, column=3, value=entry.get('urlaub_anspruch', 0))
        ws.cell(row=row_idx, column=4, value=entry.get('urlaub_uebertrag', 0))
        ws.cell(row=row_idx, column=5, value=entry.get('urlaub_verfuegbar', 0))
        ws.cell(row=row_idx, column=6, value=entry.get('urlaub_genommen', 0))
        ws.cell(row=row_idx, column=7, value=entry.get('urlaub_rest', 0))
        ws.cell(row=row_idx, column=8, value=entry.get('krankheit', 0))
        ws.cell(row=row_idx, column=9, value=entry.get('schulung', 0))
        ws.cell(row=row_idx, column=10, value=entry.get('ueberstunden', 0))
        
        # Border für alle Zellen
        for col in range(1, 11):
            ws.cell(row=row_idx, column=col).border = border
    
    # Spaltenbreite anpassen
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    
    # Speichern
    wb.save(output_path)
    # Erfolg ohne Emojis ausgeben (Windows-kompatibel)
    sys.stdout.buffer.write(f"Excel erfolgreich erstellt: {output_path}\n".encode('utf-8'))


def main():
    if len(sys.argv) != 3:
        print("FEHLER: Falsche Anzahl Parameter!", file=sys.stderr)
        print("Usage: python export_to_excel.py <input.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    # JSON lesen - WICHTIG: UTF-8 explizit setzen
    try:
        # Windows-kompatibles Öffnen mit UTF-8
        import io
        with io.open(input_file, 'r', encoding='utf-8-sig') as f:
            content = f.read()
            data = json.loads(content)
        
        # Logging ohne Emojis für Windows-Konsole - direkt auf buffer schreiben
        sys.stdout.buffer.write(f"JSON gelesen: {len(data)} Eintraege\n".encode('utf-8'))
    except Exception as e:
        # Fehler direkt auf buffer schreiben
        sys.stderr.buffer.write(f"FEHLER beim Lesen der JSON: {str(e)}\n".encode('utf-8'))
        sys.exit(1)
    
    # Excel erstellen
    try:
        create_excel(data, output_file)
    except Exception as e:
        sys.stderr.buffer.write(f"FEHLER beim Erstellen der Excel: {str(e)}\n".encode('utf-8'))
        sys.exit(1)


if __name__ == '__main__':
    main()